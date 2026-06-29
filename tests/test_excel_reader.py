"""Tests for the Excel (.xlsx) reader."""

from __future__ import annotations

from pathlib import Path

import duckdb
import pytest

from schema_seance.profile import profile
from schema_seance.readers import ExcelReaderError, UnsupportedFormatError, load
from schema_seance.readers.excel import list_sheets

openpyxl = pytest.importorskip("openpyxl")


def _build_workbook(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "people"
    ws.append(["id", "name", "email", "score"])
    ws.append([1, "Alice", "alice@example.com", 9.5])
    ws.append([2, "Bob", None, 7.25])
    ws.append([3, "Carol", "carol@example.com", None])

    ws2 = wb.create_sheet("notes")
    ws2.append(["row", "note"])
    ws2.append([1, "hello"])
    ws2.append([2, "world"])
    ws2.append([3, "again"])
    ws2.append([4, "and"])
    ws2.append([5, "again"])

    wb.create_sheet("empty")

    wb.save(path)


@pytest.fixture
def workbook(tmp_path: Path) -> Path:
    path = tmp_path / "book.xlsx"
    _build_workbook(path)
    return path


def test_load_xlsx_active_sheet(workbook: Path) -> None:
    rel = load(workbook)
    assert set(rel.columns) == {"id", "name", "email", "score"}
    assert rel.count("*").fetchone()[0] == 3


def test_load_xlsx_by_sheet_name(workbook: Path) -> None:
    rel = load(workbook, sheet="notes")
    assert list(rel.columns) == ["row", "note"]
    assert rel.count("*").fetchone()[0] == 5


def test_load_xlsx_by_sheet_index(workbook: Path) -> None:
    rel = load(workbook, sheet=1)
    assert list(rel.columns) == ["row", "note"]


def test_load_xlsx_missing_sheet(workbook: Path) -> None:
    with pytest.raises(ExcelReaderError):
        load(workbook, sheet="ghost")


def test_load_xlsx_index_out_of_range(workbook: Path) -> None:
    with pytest.raises(ExcelReaderError):
        load(workbook, sheet=99)


def test_load_xlsx_empty_sheet(workbook: Path) -> None:
    with pytest.raises(ExcelReaderError):
        load(workbook, sheet="empty")


def test_list_sheets(workbook: Path) -> None:
    sheets = list_sheets(workbook)
    names = [n for n, _ in sheets]
    assert names == ["people", "notes", "empty"]
    rows = dict(sheets)
    assert rows["people"] == 3
    assert rows["notes"] == 5
    assert rows["empty"] == 0


def test_profile_xlsx_end_to_end(workbook: Path) -> None:
    rel = load(workbook)
    report = profile(rel, path=workbook)
    by_name = {c.name: c for c in report.columns}
    assert report.rows == 3
    assert report.cols == 4
    assert by_name["email"].null_pct == pytest.approx(33.33, abs=0.01)
    assert by_name["score"].null_pct == pytest.approx(33.33, abs=0.01)


def test_load_xlsx_mixed_types(tmp_path: Path) -> None:
    from openpyxl import Workbook

    path = tmp_path / "mixed.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "mix"
    ws.append(["a", "b"])
    ws.append([1, "x"])
    ws.append(["two", 2])
    wb.save(path)

    rel = load(path)
    # DuckDB sniffs to VARCHAR when types are mixed in a column.
    assert rel.count("*").fetchone()[0] == 2


def test_load_xlsm_dispatches(tmp_path: Path) -> None:
    from openpyxl import Workbook

    path = tmp_path / "macro.xlsm"
    wb = Workbook()
    ws = wb.active
    ws.append(["k", "v"])
    ws.append(["a", 1])
    wb.save(path)
    rel = load(path)
    assert list(rel.columns) == ["k", "v"]


def test_unsupported_xls_extension(tmp_path: Path) -> None:
    bogus = tmp_path / "old.xls"
    bogus.write_text("not really xls")
    with pytest.raises(UnsupportedFormatError):
        load(bogus)


def test_load_xlsx_with_explicit_connection(workbook: Path) -> None:
    con = duckdb.connect()
    rel = load(workbook, connection=con)
    assert rel.count("*").fetchone()[0] == 3
