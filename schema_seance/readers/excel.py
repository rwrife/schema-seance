"""Excel (.xlsx/.xlsm) reader backed by openpyxl + DuckDB.

We stream rows from the active sheet (or one selected by name/index)
using openpyxl in read-only mode, materialize them to a temporary CSV,
and let DuckDB's ``read_csv_auto`` sniff types. This keeps the rest of
the pipeline format-agnostic without pulling in pandas/pyarrow.

The dependency is optional: install via ``pip install schema-seance[excel]``.
"""

from __future__ import annotations

import csv
import tempfile
from pathlib import Path
from typing import Any

import duckdb


class ExcelReaderError(ValueError):
    """Raised when an Excel file cannot be opened or a sheet is missing."""


def _require_openpyxl():
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as exc:  # pragma: no cover - exercised via test
        raise ExcelReaderError(
            "Reading Excel files requires the optional 'excel' extra. "
            "Install with: pip install 'schema-seance[excel]'"
        ) from exc
    return openpyxl


def _resolve_sheet(workbook, sheet: str | int | None) -> str:
    names = list(workbook.sheetnames)
    if not names:
        raise ExcelReaderError("Workbook contains no sheets.")
    if sheet is None:
        active = workbook.active
        return active.title if active is not None else names[0]
    if isinstance(sheet, int):
        if sheet < 0 or sheet >= len(names):
            joined = ", ".join(f"{i}={n!r}" for i, n in enumerate(names))
            raise ExcelReaderError(
                f"Sheet index {sheet} out of range (workbook has {len(names)} sheets: {joined})."
            )
        return names[sheet]
    if sheet in names:
        return sheet
    if isinstance(sheet, str) and sheet.lstrip("-").isdigit():
        return _resolve_sheet(workbook, int(sheet))
    joined = ", ".join(repr(n) for n in names)
    raise ExcelReaderError(f"Sheet {sheet!r} not found. Available: {joined}.")


def _normalize_header(raw: Any, idx: int, seen: dict[str, int]) -> str:
    name = "" if raw is None else str(raw).strip()
    if not name:
        name = f"column_{idx + 1}"
    if name in seen:
        seen[name] += 1
        return f"{name}_{seen[name]}"
    seen[name] = 0
    return name


def list_sheets(path: str | Path) -> list[tuple[str, int]]:
    """Return ``[(sheet_name, approx_row_count), ...]`` for *path*.

    Row counts are openpyxl's ``max_row`` minus 1 (header), which is
    plenty for the ``--list-sheets`` summary.
    """
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelReaderError(f"Could not open workbook {path}: {exc}") from exc
    try:
        out: list[tuple[str, int]] = []
        for name in wb.sheetnames:
            ws = wb[name]
            mr = ws.max_row or 0
            out.append((name, max(0, mr - 1)))
        return out
    finally:
        wb.close()


def _extract(
    path: str | Path,
    sheet: str | int | None,
) -> tuple[str, list[str], list[list[Any]]]:
    """Return (sheet_name, header, rows) for the chosen worksheet."""
    openpyxl = _require_openpyxl()
    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelReaderError(f"Could not open workbook {path}: {exc}") from exc
    try:
        sheet_name = _resolve_sheet(wb, sheet)
        ws = wb[sheet_name]
        header: list[str] | None = None
        rows: list[list[Any]] = []
        seen: dict[str, int] = {}
        for raw in ws.iter_rows(values_only=True):
            if header is None:
                if raw is None or all(v is None for v in raw):
                    continue
                header = [_normalize_header(v, i, seen) for i, v in enumerate(raw)]
                continue
            if raw is None or all(v is None for v in raw):
                continue
            row = list(raw)
            if len(row) < len(header):
                row += [None] * (len(header) - len(row))
            elif len(row) > len(header):
                row = row[: len(header)]
            rows.append(row)
        if header is None:
            raise ExcelReaderError(
                f"Sheet {sheet_name!r} in {path} appears to be empty (no header row)."
            )
        return sheet_name, header, rows
    finally:
        wb.close()


def read(
    path: str | Path,
    *,
    connection: duckdb.DuckDBPyConnection,
    sheet: str | int | None = None,
) -> duckdb.DuckDBPyRelation:
    """Return a DuckDB relation over a worksheet in the workbook at *path*.

    The first non-empty row is treated as the header. Fully-empty rows
    are skipped. Cell values are written to a temporary CSV which DuckDB
    sniffs for types — keeping behaviour consistent with the CSV reader.
    """
    _sheet_name, header, rows = _extract(path, sheet)

    # Write a temp CSV that DuckDB can sniff. We keep the file around for
    # the lifetime of the connection by stashing the path on the relation
    # via a no-op SELECT — but a more durable approach is to materialize
    # into a DuckDB temp table immediately.
    with tempfile.NamedTemporaryFile(
        prefix="seance_xlsx_", suffix=".csv", mode="w", delete=False, encoding="utf-8", newline=""
    ) as fh:
        writer = csv.writer(fh)
        writer.writerow(header)
        for row in rows:
            writer.writerow(["" if v is None else v for v in row])
        tmp_path = fh.name

    try:
        quoted = tmp_path.replace("'", "''")
        # Materialize into a DuckDB temp table so we can delete the CSV
        # immediately and stay self-contained.
        import uuid

        tbl = f"seance_xlsx_{uuid.uuid4().hex[:8]}"
        connection.execute(
            f'CREATE TEMP TABLE "{tbl}" AS '
            f"SELECT * FROM read_csv_auto('{quoted}', sample_size=-1, header=True)"
        )
        return connection.sql(f'SELECT * FROM "{tbl}"')
    finally:
        try:
            Path(tmp_path).unlink()
        except OSError:
            pass
